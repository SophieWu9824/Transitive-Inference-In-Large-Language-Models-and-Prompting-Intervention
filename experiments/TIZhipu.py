# python 3.12.4 Anaconda 3
# Transitive Inference of Zhipu

import __init__
import os
import openpyxl
import shutil
import time
from zhipuai import ZhipuAI
# os.chdir(os.getcwd())
mykey = 'your zhipu key'
client = ZhipuAI(api_key= mykey)

os.chdir('C:/Users/sophi/Desktop/AIPsyResearch/3_TransitiveInference/1_Experiment/Code')
# read the questions to test transitive inference
data_file_name = {'Origin':'QuestionSet.xlsx',
    'MitProb':'AblationQuestionSet_Prob.xlsx',
    'MitCoT':'AblationQuestionSet_CoT.xlsx'}
QuesName = data_file_name['Origin']
# QuesName = data_file_name['MitProb']
# QuesName = data_file_name['MitCoT']
TIQuestionSet = openpyxl.load_workbook(QuesName)
os.chdir('C:/Users/sophi/Desktop/AIPsyResearch/3_TransitiveInference/1_Experiment/Code/Zhipu')
# exp conditions
Ctxt1_chain = TIQuestionSet['studentChain']
Ctxt1_jump = TIQuestionSet['studentJump']
Ctxt2_chain = TIQuestionSet['foodChain']
Ctxt2_jump = TIQuestionSet['foodJump']
Ctxt3_chain = TIQuestionSet['employeeChain']
Ctxt3_jump = TIQuestionSet['employeeJump']


def TestTI(expCond_file,file_name,new_pos):
    answers = openpyxl.Workbook()
    zhipuTI = answers.active
    test_trials = 91
    # iteration for test
    temp_index = 0
    while test_trials:
        test_trials -= 1
        temp_index += 1
        promptTI = expCond_file['B'+str(temp_index)].value
        # get Zhipu's response
        messages = [
            {'role':'system','content':"Suppose you are a rational human thinker and reasoner. Now you need to solve some reasoning problems. You should try your best to give the most likely correct answer. You will first learn some background information which is the basis of the following reasoning problems, so you should remember it. If you are ready, I will give you the background information."},
            {'role':'user','content':promptTI}
            ]
        response = client.chat.completions.create(
            model = 'glm-4',
            messages = messages,
        )
        messages.append({'role': response.choices[0].message.role,
                'content': response.choices[0].message.content})
        answerTI = response.choices[0].message.content
        print(answerTI)
        zhipuTI.cell(row=temp_index,column=1,value=answerTI)

    # save the result
    answers.save(file_name)
    old_path = 'C:/Users/sophi/Desktop/AIPsyResearch/3_TransitiveInference/1_Experiment/Code/Zhipu/'+file_name
    target_path = 'C:/Users/sophi/Desktop/AIPsyResearch/3_TransitiveInference/1_Experiment/Data/'+new_pos
    shutil.move(old_path,target_path)

if __name__ == '__main__':
    prefixSet = {'None':'','MitProb':'Prob','MitCoT':'CoT'}
    # prefix = prefixSet['MitProb']
    prefix = prefixSet['None']
    # prefix = prefixSet['MitCoT']
    TestTI(Ctxt1_chain,prefix+'Zhipu_Ctxt1_chain.xlsx','Ctxt1Chain')
    TestTI(Ctxt1_jump,prefix+'Zhipu_Ctxt1_jump.xlsx','Ctxt1Jump')
    TestTI(Ctxt2_chain,prefix+'Zhipu_Ctxt2_chain.xlsx','Ctxt2Chain')
    TestTI(Ctxt2_jump,prefix+'Zhipu_Ctxt2_jump.xlsx','Ctxt2Jump')
    TestTI(Ctxt3_chain,prefix+'Zhipu_Ctxt3_chain.xlsx','Ctxt3Chain')
    TestTI(Ctxt3_jump,prefix+'Zhipu_Ctxt3_jump.xlsx','Ctxt3Jump')

