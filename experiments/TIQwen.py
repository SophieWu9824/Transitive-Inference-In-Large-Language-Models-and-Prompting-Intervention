
# Transitive Inference of Tongyi Qianwen

import __init__ 
import os
import openpyxl
import shutil
import random
from http import HTTPStatus # tongyi API by HTTP
import dashscope
from dashscope import Generation
from dashscope.api_entities.dashscope_response import Role

# Tongyi API
dashscope.api_key = 'your dashscope api' 
os.chdir('C:/Users/sophi/Desktop/AIPsyResearch/3_TransitiveInference/1_Experiment/Code')
# read the questions to test transitive inference
data_file_name = {'Origin':'QuestionSet.xlsx',
                  'MitProb':'AblationQuestionSet_Prob.xlsx',
                  'MitCoT':'AblationQuestionSet_CoT.xlsx'}
# QuesName = data_file_name['MitProb']
QuesName = data_file_name['MitCoT']
TIQuestionSet = openpyxl.load_workbook(QuesName)
os.chdir('C:/Users/sophi/Desktop/AIPsyResearch/3_TransitiveInference/1_Experiment/Code/Qwen')
# exp conditions
Ctxt1_chain = TIQuestionSet['studentChain']
Ctxt1_jump = TIQuestionSet['studentJump']
Ctxt2_chain = TIQuestionSet['foodChain']
Ctxt2_jump = TIQuestionSet['foodJump']
Ctxt3_chain = TIQuestionSet['employeeChain']
Ctxt3_jump = TIQuestionSet['employeeJump']

def call_with_messages(test_prompt):
    messages = [{'role':'system','content':'Suppose you are a rational human thinker and reasoner. Now you need to solve some reasoning problems. You should try your best to give the most likely correct answer. You will first learn some background information which is the basis of the following reasoning problems, so you should remember it. If you are ready, I will give you the background information.'},
                {'role':'user','content':test_prompt}]
    response = Generation.call(
        Generation.Models.qwen_turbo,
        messages = messages,
        seed =  random.randint(1,10000),
        result_format = 'message',
    )
    answer = response.output.choices[0]['message']['content']
    print(answer)
    if response.status_code == HTTPStatus.OK:
        print("/nTongYi: ", answer) 
        messages.append({'role': response.output.choices[0]['message']['role'],
                            'content': response.output.choices[0]['message']['content']})
    else:
        print('Request id: %s, Status code: %s, error code: %s, error message: %s' % (
            response.request_id, response.status_code,
            response.code, response.message
        ))
    return answer

def TestTI(expCond_file,file_name,new_pos):
    answers = openpyxl.Workbook()
    qwenTI = answers.active
    test_trials = 91
    # iteration for test
    temp_index = 0
    while test_trials:
        test_trials -= 1
        temp_index += 1
        promptTI = expCond_file['B'+str(temp_index)].value
        answerTI = call_with_messages(promptTI)
        qwenTI.cell(row=temp_index,column=1,value=answerTI)
    # save the result
    answers.save(file_name)
    old_path = 'C:/Users/sophi/Desktop/AIPsyResearch/3_TransitiveInference/1_Experiment/Code/Qwen/'+file_name
    target_path = 'C:/Users/sophi/Desktop/AIPsyResearch/3_TransitiveInference/1_Experiment/Data/'+new_pos
    shutil.move(old_path,target_path)

if __name__ == '__main__':
    prefixSet = {'None':'','MitProb':'Prob','MitCoT':'CoT'}
    # prefix = prefixSet['MitProb']
    # prefix = prefixSet['None']
    prefix = prefixSet['MitCoT']
    TestTI(Ctxt1_chain,prefix+'Qwen_Ctxt1_chain.xlsx','Ctxt1Chain')
    TestTI(Ctxt1_jump,prefix+'Qwen_Ctxt1_jump.xlsx','Ctxt1Jump')
    TestTI(Ctxt2_chain,prefix+'Qwen_Ctxt2_chain.xlsx','Ctxt2Chain')
    TestTI(Ctxt2_jump,prefix+'Qwen_Ctxt2_jump.xlsx','Ctxt2Jump')
    TestTI(Ctxt3_chain,prefix+'Qwen_Ctxt3_chain.xlsx','Ctxt3Chain')
    TestTI(Ctxt3_jump,prefix+'Qwen_Ctxt3_jump.xlsx','Ctxt3Jump')
