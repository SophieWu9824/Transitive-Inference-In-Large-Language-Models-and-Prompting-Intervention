# Transitive Inference of GPT
# Anaconda3 python 3.12.4

import os
import openpyxl
import shutil
import requests
import time
import requests.adapters

os.chdir('C:/Users/sophi/Desktop/AIPsyResearch/3_TransitiveInference/1_Experiment/Code')
# read the questions to test transitive inference
data_file_name = {'Origin':'QuestionSet.xlsx',
    'MitProb':'AblationQuestionSet_Prob.xlsx',
    'MitCoT':'AblationQuestionSet_CoT.xlsx'}
# QuesName = data_file_name['Origin']
# QuesName = data_file_name['MitProb']
QuesName = data_file_name['MitCoT']
TIQuestionSet = openpyxl.load_workbook(QuesName)
os.chdir('C:/Users/sophi/Desktop/AIPsyResearch/3_TransitiveInference/1_Experiment/Code/OpenAI')
# exp conditions
Ctxt1_chain = TIQuestionSet['studentChain']
Ctxt1_jump = TIQuestionSet['studentJump']
Ctxt2_chain = TIQuestionSet['foodChain']
Ctxt2_jump = TIQuestionSet['foodJump']
Ctxt3_chain = TIQuestionSet['employeeChain']
Ctxt3_jump = TIQuestionSet['employeeJump']

def TestTI(expCond_file,file_name,new_pos):
    answers = openpyxl.Workbook()
    gptTI = answers.active
    test_trials = 91
    temp_index = 0
    messages = [{"role": "system",
                "content": "Suppose you are a rational human thinker and reasoner. Now you need to solve some reasoning problems. You should try your best to give the most likely correct answer. You will first learn some background information which is the basis of the following reasoning problems, so you should remember it. If you are ready, I will give you the background information."
                }]
    while test_trials:
        test_trials -= 1
        temp_index += 1
        promptTI = expCond_file['B'+str(temp_index)].value
        messages.append({"role": "user","content": promptTI})
        # answer = ViaRequest(promptTI)
       
        url="https://api.openai.com/v1/chat/completions" 
        # OPENAI_API_KEY="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"
        OPENAI_API_KEY="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"
        header={"Content-Type": "application/json","Authorization": "Bearer " +OPENAI_API_KEY}
        data={
            "model": "gpt-3.5-turbo", # gpt-3.5-turbo, gpt-4
            "messages": messages,
            "temperature":0,
            "stream":False,    
        }
        try:
            requests.adapters.DEFAULT_RETRIES = 15
            s = requests.session()
            s.keep_alive = False
            response=requests.post(url=url,headers=header,json=data).json()
            # print("OpenAI: ",response)
            answer = response['choices'][0]['message']['content']
            messages.append({'role': response['choices'][0]['message']['role'],
                    'content': response['choices'][0]['message']['content']})
            # time.sleep(1)
        except Exception:
            print('Error!')
            time.sleep(0.5)
            answer = 'Error'
        print(gpt_name+": ",answer)
        gptTI.cell(row=temp_index,column=1,value=answer)
    
    # save the result
    answers.save(file_name)
    old_path = 'C:/Users/sophi/Desktop/AIPsyResearch/3_TransitiveInference/1_Experiment/Code/OpenAI/'+file_name
    target_path = 'C:/Users/sophi/Desktop/AIPsyResearch/3_TransitiveInference/1_Experiment/Data/'+new_pos
    shutil.move(old_path,target_path)   

if __name__ == "__main__":
    prefixSet = {'None':'','MitProb':'Prob','MitCoT':'CoT'}
    # prefix = prefixSet['None']
    # prefix = prefixSet['MitProb']
    prefix = prefixSet['MitCoT']
    gpt_name = "gpt3.5-turbo" 
    # gpt_name = "gpt4"
    # TestTI(Ctxt1_chain,prefix+gpt_name+'_Ctxt1_chain.xlsx','Ctxt1Chain')
    # TestTI(Ctxt1_jump,prefix+gpt_name+'_Ctxt1_jump.xlsx','Ctxt1Jump')
    # TestTI(Ctxt2_chain,prefix+gpt_name+'_Ctxt2_chain.xlsx','Ctxt2Chain')
    TestTI(Ctxt2_jump,prefix+gpt_name+'_Ctxt2_jump.xlsx','Ctxt2Jump')
    TestTI(Ctxt3_chain,prefix+gpt_name+'_Ctxt3_chain.xlsx','Ctxt3Chain')
    TestTI(Ctxt3_jump,prefix+gpt_name+'_Ctxt3_jump.xlsx','Ctxt3Jump')
