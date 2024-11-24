# Transitive Inference of Llama3
# python 3.8

import os
import openpyxl
import shutil
import time
from openai import OpenAI

os.chdir('C:/Users/sophi/Desktop/AIPsyResearch/3_TransitiveInference/1_Experiment/Code')
# read the questions to test transitive inference
data_file_name = {'Origin':'QuestionSet.xlsx',
    'MitProb':'AblationQuestionSet_Prob.xlsx',
    'MitCoT':'AblationQuestionSet_CoT.xlsx'}
# QuesName = data_file_name['Origin']
QuesName = data_file_name['MitProb']
# QuesName = data_file_name['MitCoT']
TIQuestionSet = openpyxl.load_workbook(QuesName)
os.chdir('C:/Users/sophi/Desktop/AIPsyResearch/3_TransitiveInference/1_Experiment/Code/Llama')
# exp conditions
Ctxt1_chain = TIQuestionSet['studentChain']
Ctxt1_jump = TIQuestionSet['studentJump']
Ctxt2_chain = TIQuestionSet['foodChain']
Ctxt2_jump = TIQuestionSet['foodJump']
Ctxt3_chain = TIQuestionSet['employeeChain']
Ctxt3_jump = TIQuestionSet['employeeJump']

def TestTI(expCond_file,file_name,new_pos):
    # nvidia_api_key = os.getenv("NVIDIA_API_KEY")
    nvidia_api_key = os.getenv("NVIDIA_API_KEY2")
    client = OpenAI(
        base_url= "https://integrate.api.nvidia.com/v1",
        api_key = nvidia_api_key
    )
    answers = openpyxl.Workbook()
    gptTI = answers.active
    test_trials = 91
    temp_index = 0
    messages = [{"role": "system",
                "content": "Suppose you are a rational human thinker and reasoner. Now you need to solve some reasoning problems. You should try your best to give the most likely correct answer. You will first learn some background information which is the basis of the following reasoning problems, so you should remember it. If you are ready, I will give you the background information."
                }]
    while test_trials:
        test_trials -= 1
        temp_index += 1
        promptTI = expCond_file['B'+str(temp_index)].value
        messages.append({"role": "user","content": promptTI})
        response = client.chat.completions.create(
            model = "meta/llama3-8b-instruct",
            messages = messages,
            temperature = 0, # 0
            stream = False
        )
        answer = response.choices[0].message.content
        # retain memory
        messages.append({'role':response.choices[0].message.role,
                         'content':response.choices[0].message.content})
        print(response.model+": ",answer)
        gptTI.cell(row=temp_index,column=1,value=answer)
        
    # save the result
    answers.save(file_name)
    old_path = 'C:/Users/sophi/Desktop/AIPsyResearch/3_TransitiveInference/1_Experiment/Code/Llama/'+file_name
    target_path = 'C:/Users/sophi/Desktop/AIPsyResearch/3_TransitiveInference/1_Experiment/Data/'+new_pos
    shutil.move(old_path,target_path)   

if __name__ == "__main__":
    prefixSet = {'None':'','MitProb':'Prob','MitCoT':'CoT'}
    # prefix = prefixSet['None']
    prefix = prefixSet['MitProb']
    # prefix = prefixSet['MitCoT']
    model_name = "llama3-8b" 
    # TestTI(Ctxt1_chain,prefix+model_name+'_Ctxt1_chain.xlsx','Ctxt1Chain')
    # TestTI(Ctxt1_jump,prefix+model_name+'_Ctxt1_jump.xlsx','Ctxt1Jump')
    # TestTI(Ctxt2_chain,prefix+model_name+'_Ctxt2_chain.xlsx','Ctxt2Chain')
    # TestTI(Ctxt2_jump,prefix+model_name+'_Ctxt2_jump.xlsx','Ctxt2Jump')
    # TestTI(Ctxt3_chain,prefix+model_name+'_Ctxt3_chain.xlsx','Ctxt3Chain')
    TestTI(Ctxt3_jump,prefix+model_name+'_Ctxt3_jump.xlsx','Ctxt3Jump')
